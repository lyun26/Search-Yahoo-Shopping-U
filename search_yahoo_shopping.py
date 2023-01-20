import os
import argparse
import openpyxl

from yahoo_api import searchItems

from functools import wraps
import time

def stop_watch(func):
    @wraps(func)
    def wrapper(*args, **kargs):
        start = time.perf_counter()
        result = func(*args, **kargs)
        elapsed_time = time.perf_counter() - start
        print('Time: {} s ({} min)'.format(elapsed_time, elapsed_time//60))
        return result
    return wrapper


def option():
    parser = argparse.ArgumentParser(
        description=''
    )
    parser.add_argument('-v', '--version',
                        help='version information',
                        action='version',
                        version='%(prog)s 0.1'
                        )
    parser.add_argument('-o', '--output',
                        help='specifiy output folder',
                        action='store',
                        default='out'
                        )
    parser.add_argument('-k', '--keyword-file',
                        help='input keyword file',
                        default='keyword.xlsx')
    parser.add_argument('-a', '--appid_file',
                        help='appid',
                        default='appid.xlsx'
                        )
    parser.add_argument('-m', '--max_number',
                        help='max number of search result',
                        type=int,
                        default=1000000)
    parser.add_argument('--only-search-store',
                        help='ONLY Search for stores related the keywords, output shops_all.xlsx',
                        action='store_true',
                        default=False)
    parser.add_argument('--only-search-item',
                        help='ONLY Search for items of stores related shops_all.xlsx specified',
                        action='store_true',
                        default=False)
    parser.add_argument('--shops-all-path',
                        help='specify shops_all.xlsx path to run --only-search-store option',
                        default='shops_all.xlsx'
                        )

    return parser.parse_args()

def load_xlsx_cells(xlsx:str):
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[wb.sheetnames[0]]

    values = []
    clm = 0
    for r in range(ws.max_row):
        cell = ws['A'+str(r+1)]
        values.append(cell.value)
    return values


def load_keywords(keyword_file:str):
    keywords = load_xlsx_cells(keyword_file)
    return keywords


def load_appids(appid_file: str):
    t = load_xlsx_cells(appid_file)
    appids = list(map(lambda x: x.strip(), t))
    return appids


def load_limit_conf(keyword_file):
    wb = openpyxl.load_workbook(keyword_file)
    ws = wb[wb.sheetnames[0]]

    max_items_per_xlsx = int(ws['B1'].value)
    max_shops = int(ws['C1'].value)
    min_items_per_shop = int(ws['D1'].value)
    return (max_items_per_xlsx, max_shops, min_items_per_shop)


@stop_watch
def main():
    args = option()
    print(args)
    keywords = load_keywords(args.keyword_file)
    appids = load_appids(args.appid_file)
    max_items_per_xlsx, max_shops, min_item_per_shop = load_limit_conf(args.keyword_file)

    os.makedirs(args.output, exist_ok=True)
    searchItem = searchItems(keywords, appids, args.output, args.max_number,
                                max_items_per_xlsx, max_shops, min_item_per_shop)

    
    if args.only_search_store:
        searchItem.run_only_search_stores()  # this code works
    elif args.only_search_item:
        searchItem.run_only_search_items(args.shops_all_path)
    else:
        searchItem.run()
    return 0


if __name__ == '__main__':
    exit(main())
