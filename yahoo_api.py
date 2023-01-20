import os
import glob
import time
import httpx
import openpyxl

import queue
import threading

from yahoo_scraping import scraipe_yahoo_shopsite


itemSearch_ep = 'https://shopping.yahooapis.jp/ShoppingWebService/V3/itemSearch'
max_rpm = 30 # max requests per minute
MAX_RETURNED_RESULTS = 1000
Is_SearchShop_alive = True


def create_query_params(appid, get_results, add_params={}):
    params = {'appid': appid,
                'results': get_results}
    for k, v in add_params.items():
        params[k] = v # 同一キーは上書き
    return params


def retry_request(client: httpx.Client, extra_appids=[]):
    MAX_TMR_NUM = 5 # 429: Too Many Requests を連続で返却されたときの許容数
    tmr_cnt = 1
    while tmr_cnt < MAX_TMR_NUM:
        #print('...retry cnt:', tmr_cnt)
        time.sleep(60) # 1分のチェックあたりまで待機
        #print('retry...')
        try:
            r = client.get(url=itemSearch_ep)
            if r.status_code == 200:
                tmr_cnt = 0
                return r
        except:
            pass
        #print('[-] Error', r.status_code, r.text)
        tmr_cnt += 1
    print('[-] Error', 'OVER MAX retry number')
    return None


def recieve_response(r:httpx.Response, client: httpx.Client, extra_appids=[]):
    if r.status_code == 200:
        rData = r.json()
        #print('[+]Check results:', rData['totalResultsAvailable'], rData['totalResultsReturned'], rData['firstResultsPosition'], client.params)
        return r.json()
    elif r.status_code == 429:
        # Too Many Requests
        #print("[-]Error: Too Many Requests.", client.params)
        r = retry_request(client, extra_appids)
        if not r:
            print('[-] Too Many Request')
            return None
        return r.json()
    else:
        print('Status Code:', r.status_code, ', Text:', r.text, client.params)
        r = retry_request(client, extra_appids)
        if not r:
            return None
        return r.json()


def get_request(params: dict, extra_appids=[]):
    client = httpx.Client(params=params)
    try_cnt = 0
    MAX_RETRY_NUM = 5
    while try_cnt < MAX_RETRY_NUM:
        try:
            r = client.get(url=itemSearch_ep)
            rData = recieve_response(r, client, extra_appids)
            return rData
        except:
            pass
        try_cnt += 1
    return None


def create_price_range(params: dict, price_start=1, max_item_number:int=10000):
    price_range = []
    pStart = price_start
    local_params = {}
    for k, v in params.items():
        local_params[k] = v
    # 降順で最大金額を取得
    local_params['sort'] = '-price'
    client = httpx.Client(params=local_params)
    r = client.get(url=itemSearch_ep)
    rData = recieve_response(r, client)
    pStart, pEnd = price_start, rData['hits'][0]['price']
    totalResults = rData['totalResultsAvailable']
    tmp_list = [[totalResults, pStart, pEnd]]
    if totalResults < MAX_RETURNED_RESULTS:
        return tmp_list

    # 検索結果がMAX_RETURNED_RESULTSに収まる価格範囲のリストを作成
    local_params['sort'] = '+price'
    pFrom, pTo = pStart, pEnd
    availableSum = 0
    while tmp_list and availableSum < max_item_number:
        _, pFrom, pTo = tmp_list[-1]
        local_params['price_from'] = pFrom
        middle = (pFrom + pTo) // 2
        local_params['price_to'] = middle
        rData = get_request(params=local_params)
        if not rData:
            break
        availableResults = rData['totalResultsAvailable']
        if availableResults == 0:
            tmp_list[-1][1] = middle + 1
        elif availableResults < MAX_RETURNED_RESULTS:
            price_range.append([availableResults, pFrom, middle])
            tmp_list[-1][0] -= availableResults
            tmp_list[-1][1] = middle + 1
            if tmp_list[-1][0] == 0:
                del tmp_list[-1]
            elif tmp_list[-1][0] < MAX_RETURNED_RESULTS:
                price_range.append(tmp_list.pop())
                if price_range[-1][0] >= 1000:
                    availableSum += 1000
                else:
                    availableSum += price_range[-1][0]
        else:
            new_elm = [availableResults, pFrom, middle]
            tmp_list[-1][0] -= new_elm[0]
            tmp_list[-1][1] = middle + 1
            if tmp_list[-1][0] == 0:
                del tmp_list[-1]
            elif tmp_list[-1][0] < MAX_RETURNED_RESULTS \
                    or tmp_list[-1][1] == tmp_list[-1][2]:
                price_range.append(tmp_list.pop())
                if price_range[-1][0] >= 1000:
                    availableSum += 1000
                else:
                    availableSum += price_range[-1][0]
            if pFrom == middle:
                #print('[+] cut results', middle, availableResults)
                price_range.append(new_elm)
                if new_elm[0] >= 1000:
                    availableSum += 1000
                else:
                    availableSum += new_elm[0]
            else:
                tmp_list.append(new_elm)

    if tmp_list:
        price_range += tmp_list        
    price_range.sort(key=lambda x: x[1])
    return price_range


class SearchItemOfShop(threading.Thread):
    def __init__(self, appid, max_items_number:int, shop_queue: queue.Queue, extra_appids:list):
        super(SearchItemOfShop, self).__init__()
        self.appid = appid
        self.extra_appids = extra_appids
        self.max_items_number = max_items_number
        self.get_results = 100
        self.shop_queue = shop_queue
        self.items = []

    def is_yahoo_shop_store(sel, shop_url):
        return shop_url.startswith('https://store.shopping.yahoo.co.jp/')


    def scraipe_yahoo_shop(self, shop, pFrom, pTo, availableResults):
        r = scraipe_yahoo_shopsite(shop['url'], pFrom, pTo, availableResults)
        return r


    def request_yahoo_api(self, shop, pFrom, pTo, availableResults, checkedResults):
        hits = []
        params ={'appid': self.appid,
                    'results': self.get_results,
                    'seller_id': shop['seller_id'],
                    'sort': '+price'}
        params['start'] = 1
        params['price_from'] = pFrom
        params['price_to'] = pTo
        while params['start'] < availableResults \
                and params['start'] < MAX_RETURNED_RESULTS\
                and checkedResults + len(hits) < self.max_items_number:
            rdata = get_request(params=params)
            if not rdata:
                return False
            hits += rdata['hits']
            params['start'] += self.get_results
            if params['start'] + self.get_results > MAX_RETURNED_RESULTS:
                params['results'] = MAX_RETURNED_RESULTS - params['start']
        return hits


    def run(self):
        '''
        ショップごとの商品を安い方から規定数保存
        '''
        global Is_SearchShop_alive
        print('[{}] Start Search Item Thread'.format(self.native_id))
        while True:
            if self.shop_queue.empty():
                if Is_SearchShop_alive:
                    time.sleep(10)
                    continue
                else:
                    break

            start_time = time.perf_counter()
            shop = self.shop_queue.get()
            xlsx_fname = os.path.join(shop['shop_folder'], shop['shop_fname'])
            xlsx_wb = openpyxl.Workbook()
            xlsx_ws = xlsx_wb.active

            # 店舗の商品を検索してxlsxへ保存
            # 商品名, 値段, 店舗のURL
            start_time_pr = time.perf_counter()
            price_range = create_price_range({'appid': self.appid,
                                            'results': self.get_results,
                                            'seller_id': shop['seller_id']},
                                            price_start=1,
                                            max_item_number=self.max_items_number)
            elapsed_time_pr = time.perf_counter() - start_time_pr
            #print('[{}] {} finish. time: {}s ({}min) {}'.format(self.native_id, 'create_price_range', elapsed_time_pr, elapsed_time_pr//60, price_range))
            checkedResults = 0

            # 同一(商品名、価格)の重複削除。商品名と価格をキーに要素作成。値は使わないのでTrue
            # エクセルへの出力前に重複確認
            # items:= { ("NameA", priceA): True, ("NameA", priceB): True, ("NameB", priceA): True }
            items = {}
            for pr in price_range:
                availableResults, pFrom, pTo = pr
                if availableResults >= MAX_RETURNED_RESULTS:
                        if self.is_yahoo_shop_store(shop['url']):
                            hits = self.scraipe_yahoo_shop(shop, pFrom, pTo, availableResults)
                        else:
                            hits = self.request_yahoo_api(shop, pFrom, pTo, availableResults, checkedResults)
                else:
                    hits = self.request_yahoo_api(shop, pFrom, pTo, availableResults, checkedResults)

                for hit in hits:
                    name = hit['name']
                    price = hit['price']
                    if (name, price) in items.keys():
                        continue

                    checkedResults += 1
                    items[(name, price)] = True
                    xlsx_ws['A' + str(checkedResults)] = name
                    xlsx_ws['B' + str(checkedResults)] = price
                    xlsx_ws['C' + str(checkedResults)] = shop['url']
                    if checkedResults >= self.max_items_number:
                        break
                xlsx_wb.save(xlsx_fname)
                if checkedResults >= self.max_items_number:
                    break
            elapsed_time = time.perf_counter() - start_time
            items.clear()
            print('[{}] {} finish. num: {}, time: {}s ({}min)'.format(self.native_id, shop['name'], checkedResults, elapsed_time, elapsed_time//60))
        print('[{}] Finish Search Item Thread'.format(self.native_id))


class SearchShops(threading.Thread):
    def __init__(self, appid, max_number:int, max_shops:int, min_item_per_shop:int, rData, keyword, keyword_folder, shop_queue:queue.Queue):
        super(SearchShops, self).__init__()
        self.appid = appid
        self.rData = rData
        self.get_results = 100
        self.shops = {}
        self.max_number = max_number
        self.max_shops = max_shops
        self.min_item_per_shop = min_item_per_shop
        self.target_shop_cnt = 0
        self.keyword = keyword
        self.keyword_folder = keyword_folder
        self.shop_queue = shop_queue # SearchItemsOfShop にショップ情報を渡すキュー

        self.xlsx_fname = os.path.join(self.keyword_folder, 'shops.xlsx')
        self.xlsx_wb = openpyxl.Workbook()
        self.xlsx_ws = self.xlsx_wb.active
        # 最初のデータのseller情報をショップ情報のスクレイプングスレッドへ投げておく
        for hit in rData['hits']:
            if len(self.shops) >= self.max_shops:
                break
            seller = hit['seller']
            seller_id = seller['sellerId']
            if seller_id not in self.shops.keys():
                self.save_seller_info(seller)


    def save_seller_info(self, hit_seller):
        seller_name = hit_seller['name']
        seller_id = hit_seller['sellerId']
        seller_url = hit_seller['url']
        shop_fname = seller_name + '_' + seller_id + '.xlsx'
        shop_folder = os.path.join(self.keyword_folder, 'shop')

        self.shops[seller_id] = {
            'seller_id': seller_id,
            'name': seller_name,
            'url': seller_url,
            'keyword': self.keyword,
            'shop_folder': shop_folder,
            'shop_fname': shop_fname}

        # 検索をかける店舗の商品数の下限チェック
        params = create_query_params(self.appid, self.get_results,
                                    {'seller_id': seller_id})
        rdata = get_request(params=params)
        if not rdata:
            return
        totalResults = rdata['totalResultsAvailable']
        if totalResults < self.min_item_per_shop:
            return

        # 店舗の商品情報を商品検索スレッドに登録、店舗情報をエクセルに出力
        self.target_shop_cnt += 1
        print('[{}] Item Count for {}, total results: {}'.format(self.target_shop_cnt, seller_name, totalResults))
        self.shop_queue.put(self.shops[seller_id])
        # xlsxに追記
        # url, 店舗名, 店舗ID
        self.xlsx_ws['A' + str(self.target_shop_cnt)] = seller_url
        self.xlsx_ws['B' + str(self.target_shop_cnt)] = seller_name
        self.xlsx_ws['C' + str(self.target_shop_cnt)] = seller_id
        self.xlsx_wb.save(self.xlsx_fname)


    def run(self):
        '''
        Yahoo! Shoppingの検索結果は、1000件までしか取得できない (start + results <= 1000)
        なので、検索結果が1000件に絞られるように値段幅を変更して店舗を全件取得
        '''
        totalResults = self.rData['totalResultsAvailable']
        checkedResults = 0

        '''
        price_range = create_price_range(params={'appid': self.appid,
                                                'results': self.get_results,
                                                'query': self.keyword})
        '''
        if totalResults >= MAX_RETURNED_RESULTS:
            params_minus ={'appid': self.appid,
                                'results': self.get_results,
                                'sort': '-price',
                                'query': self.keyword}
            r_minusData = get_request(params=params_minus)
            if not r_minusData:
                return False
            pStart, pEnd = self.rData['hits'][0]['price'], r_minusData['hits'][0]['price']
        else:
            pStart, pEnd = self.rData['hits'][0]['price'], self.rData['hits'][-1]['price']

        pFrom, pTo = pStart, pEnd
        while checkedResults < totalResults:
            availableResults = 0
            resultsSave = [] # [(availableResults, pFrom, pTo), ...]
            params = create_query_params(self.appid,
                                        self.get_results,
                                        {'query': self.rData['request']['query'],
                                        'sort': '+price',
                                        'start': 1})
            if totalResults >= 1000:
                while True:
                    params['price_from'] = pFrom
                    params['price_to'] = pTo
                    rdata = get_request(params=params)
                    if not rdata:
                        return False
                    availableResults = rdata['totalResultsAvailable']

                    if availableResults > 0:
                        resultsSave.append((availableResults, pFrom, pTo))

                    if availableResults == 0:
                        if resultsSave:
                            pFrom, pTo = pTo + 1, resultsSave[-1][2]
                        else:
                            pFrom, pTo = pTo + 1, pEnd
                    elif availableResults < 1000:
                        break
                    else:
                        pFrom = rdata['hits'][0]['price']
                        pTo = (pFrom + pTo) // 2

                    if pFrom == params['price_to']:
                        #print('[{}] cut same price. {}-{}.'.format(self.native_id, pFrom, pTo))
                        break
                    elif pTo <= pFrom:
                        pTo = pFrom
            else:
                availableResults = totalResults

            pFrom, pTo = pTo + 1, pEnd
            if resultsSave:
                rs_reverse = resultsSave[::-1]
                for rs in rs_reverse:
                    if rs[0] - availableResults > 0:
                        #print('[+]   pTo=', rs[2], rs)
                        pTo = rs[2]
                        break
                    else:
                        #print('[+] pFrom=', rs[2]+1, rs)
                        pFrom = rs[2] + 1

            # save shops
            while params['start'] < availableResults \
                    and params['start'] < MAX_RETURNED_RESULTS \
                    and self.target_shop_cnt < self.max_shops:
                #print(params)
                rdata = get_request(params=params)
                if not rdata:
                    return False
                rReturned = rdata['totalResultsReturned']
                hits = rdata['hits']
                for hit in hits:
                    if self.target_shop_cnt >= self.max_shops:
                        break
                    seller_id = hit['seller']['sellerId']
                    if seller_id not in self.shops.keys():
                        self.save_seller_info(hit['seller'])

                params['start'] += self.get_results
                if params['start'] + self.get_results > MAX_RETURNED_RESULTS:
                    params['results'] = MAX_RETURNED_RESULTS - params['start']
                    #print('set results', params['results'])
                checkedResults += rReturned
            params['results'] = self.get_results
            if pFrom > pEnd or self.target_shop_cnt >= self.max_shops:
                break
        print('[{}] search shops finish: {}, {}'.format(self.name, self.keyword, len(self.shops)))
        return True


class searchItems:
    MAX_RETURNED_RESULTS = 1000 # Yahoo APIの制限。取得できる検索結果の上限
    def __init__(self, keywords:list, appids:list,
                    output:str, max_number:int,
                    max_items_per_xlsx:int, max_shops:int,
                    min_item_per_shop:int):
        self.keywords = keywords
        self.appids = appids
        self.appid = self.appids[0]
        self.output = output
        self.output_fname = 'out'
        self.max_number = max_number
        self.get_results = 100
        self.shops = {}
        self.results = {}
        self.tmr_cnt = 0 # 429: Too Many Requests を返却された回数
        self.max_items_per_xlsx = max_items_per_xlsx
        self.max_shops = max_shops
        self.min_item_per_shop = min_item_per_shop


    def merge_shops(self):
        output_xlsx = os.path.join(self.output_folder, 'shops_all.xlsx')
        output_wb = openpyxl.Workbook(write_only=True)
        output_ws = output_wb.create_sheet()

        shops = {}
        shops_queue = queue.Queue()
        for keyword in self.keywords:
            xlsx_path = os.path.join(self.output_folder, keyword, 'shops.xlsx')
            read_wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            read_ws = read_wb.active
            for row in read_ws.iter_rows():
                url = row[0].value
                name = row[1].value
                id = row[2].value
                if url not in shops.keys():
                    shops[url] = {"name": name, "seller_id": id}
                    shops_queue.put((url, name, id))
            read_wb.close()

        while not shops_queue.empty():
            shop_url, shop_name, seller_id = shops_queue.get()
            output_ws.append(row=[shop_url, shop_name, seller_id])
        output_wb.save(output_xlsx)
        return


    def merge_items(self):
        def get_new_output_ws(xlsx_cnt:int):
            output_xlsx = os.path.join(self.output_folder,
                                'items_all_{}.xlsx'.format(xlsx_cnt))
            output_wb = openpyxl.Workbook(write_only=True)
            output_ws = output_wb.create_sheet()
            return (output_wb, output_ws, output_xlsx)

        xlsx_cnt = 1
        write_rows_cnt = 0
        output_wb, output_ws, output_xlsx = get_new_output_ws(xlsx_cnt)
        for keyword in self.keywords:
            keyword_folder = os.path.join(self.output_folder, keyword, 'shop')
            xlsx_files = glob.glob(keyword_folder + '/*.xlsx')
            for xlsx_file in xlsx_files:
                read_wb = openpyxl.load_workbook(xlsx_file, read_only=True)
                read_ws = read_wb.active
                for row in read_ws.iter_rows():
                    name = row[0].value
                    price = row[1].value
                    store_url = row[2].value
                    output_ws.append(row=[name, price, store_url])
                    write_rows_cnt += 1
                    if write_rows_cnt >= self.max_items_per_xlsx:
                        output_wb.save(output_xlsx)
                        xlsx_cnt += 1
                        write_rows_cnt = 0
                        output_wb, output_ws, output_xlsx = get_new_output_ws(xlsx_cnt)
                read_wb.close()
        output_wb.save(output_xlsx)


    def run(self):
        global Is_SearchShop_alive
        shop_queue = queue.Queue()

        # save items of each shop
        searchItemsOfShop = SearchItemOfShop(self.appids[-1], self.max_number,
                                            shop_queue, self.appids[1:-1])
        searchItemsOfShop.start()
        self.output_folder = os.path.join(self.output, time.strftime('%Y%m%d_%H%M'))
        for keyword in self.keywords:
            keyword_folder = os.path.join(self.output_folder, keyword)
            try:
                os.makedirs(keyword_folder, exist_ok=True)
                os.makedirs(os.path.join(keyword_folder, 'shop'), exist_ok=True)
            except Exception as e:
                print(e)
                return -1

            params = create_query_params(
                                self.appid, self.get_results,
                                {'query': keyword,
                                'sort': '+price'})
            rData = get_request(params=params)
            if not rData:
                return -1
            print('[+] Success: Initial Request')
            print('[+] keyword: {}, search result: {}'.format(keyword, rData['totalResultsAvailable']))

            # save shops
            searchShopsThread = SearchShops(self.appids[0],
                                                self.max_number, self.max_shops,
                                                self.min_item_per_shop,
                                                rData, keyword,
                                                keyword_folder, shop_queue)

            searchShopsThread.start()

            searchShopsThread.join()
            print('[0] search keyword finish:', keyword)

        Is_SearchShop_alive = False
        # ショップ情報のxlsxを1つにマージ。重複排除
        self.merge_shops()

        # 商品検索完了するまで待ち
        searchItemsOfShop.join()
        # 各ショップの商品情報をxlsxを1つにマージ。
        # ただし、1つのxlsxの最大行数はdefault 50万行まで
        # 超えたら、items_all_{number}.xlsxの形式で連番で作成
        self.merge_items()
        return 0


    def run_only_search_stores(self):
        global Is_SearchShop_alive
        shop_queue = queue.Queue()  # 作って通知いれるだけ。処理するスレッドなし

        self.output_folder = os.path.join(self.output, time.strftime('%Y%m%d_%H%M'))
        for keyword in self.keywords:
            keyword_folder = os.path.join(self.output_folder, keyword)
            try:
                os.makedirs(keyword_folder, exist_ok=True)
            except Exception as e:
                print(e)
                return -1

            params = create_query_params(
                                self.appid, self.get_results,
                                {'query': keyword,
                                'sort': '+price'})
            rData = get_request(params=params)
            if not rData:
                return -1
            print('[+] Success: Initial Request')
            print('[+] keyword: {}, search result: {}'.format(keyword, rData['totalResultsAvailable']))

            # save shops
            searchShopsThread = SearchShops(self.appids[0],
                                                self.max_number, self.max_shops,
                                                rData, keyword,
                                                keyword_folder, shop_queue)
            searchShopsThread.start()

            searchShopsThread.join()
            print('[0] search keyword finish:', keyword)

        # ショップ情報のxlsxを1つにマージ。重複排除
        self.merge_shops()
        return 0


    def run_only_search_items(self, shops_all_path):
        global Is_SearchShop_alive
        shop_queue = queue.Queue()

        if not os.path.exists(shops_all_path):
            print('[-] Error: File Not Found {}'.format(shops_all_path))
            return -1

        # save items of each shop
        searchItemsOfShop = SearchItemOfShop(self.appids[-1], self.max_number,
                                            shop_queue, self.appids[1:-1])
        searchItemsOfShop.start()

        self.output_folder = os.path.join(self.output, time.strftime('%Y%m%d_%H%M'))
        shop_folder = os.path.join(self.output_folder, 'shop')
        try:
            os.makedirs(shop_folder, exist_ok=True)
        except Exception as e:
            print(e)
            return -1

        read_wb = openpyxl.load_workbook(shops_all_path, read_only=True)
        read_ws = read_wb.active
        shops = {}
        for row in read_ws.iter_rows():
            url = row[0].value
            name = row[1].value
            id = row[2].value
            shop_fname = name + '_' + id + '.xlsx'
            shops[id] = {'seller_id': id,
                            'name': name,
                            'url': url,
                            'shop_folder': shop_folder,
                            'shop_fname': shop_fname}
            shop_queue.put(shops[id])

        Is_SearchShop_alive = False
        searchItemsOfShop.join()

        self.keywords = [""]
        self.merge_items()
        return 0
